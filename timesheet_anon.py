#! python3
# Timesheet Timesaver #
# Edit your times in, out and your lunch #

# Keep your times in double quotes
timein = "7:30"
lunchout = "12:00"
lunchin = "13:00"
timeout = "16:00"

import sys
import time
import getpass
import datetime

# Pull username
username = getpass.getuser()

# Edit the below to match user# to usernames of those in your team (check the spreadsheet for reference)
if username == "user1":
    MyRangeStart = 144
    MyRangeEnd = 154
elif username == "user2":
    MyRangeStart = 59
    MyRangeEnd = 69
elif username == "user3":
    MyRangeStart = 93
    MyRangeEnd = 103
else:
    print("You are not the right person!")
    sys.exit()
    
# This script should be run on Fridays, you wouldn't be logging time you haven't worked yet... Would you?
today = datetime.date.today()
weekday = today.weekday()
if weekday == 0:
    print("You can't run this script on a Monday!")
    time.sleep(3)
    sys.exit()
else:
    continue

    
# Import modules for Excel editing.
from openpyxl import load_workbook
from openpyxl.styles import Alignment

# Import module for windows filepathing.
from pathlib import Path

# This grabs the date of the last Monday that passed.
last_monday = today - datetime.timedelta(days=today.weekday())
monday = (last_monday.strftime('%d%m%Y') + ".xlsx")

# Define the folder path and filename, I think i could have just used an r string here...
folder = Path("S:/Shared/IT Team/Attendance/IT/")
filename = folder / monday

wb = load_workbook(filename)
# grab the active worksheet.
ws = wb.active

# Writing time data to the specified cells, then matching the formatting.

for cells in range(MyRangeStart,MyRangeEnd,2):
    morning = f"B{cells}" 
    ws[morning] = timein
    currentcell = ws[morning]
    currentcell.alignment = Alignment(horizontal='right')
    
    lunch = f"C{cells}"
    ws[lunch] = lunchout
    currentcell = ws[lunch]
    currentcell.alignment = Alignment(horizontal='right')
    
    lunch2 = f"D{cells}"
    ws[lunch2] = lunchin
    currentcell = ws[lunch2]
    currentcell.alignment = Alignment(horizontal='right')
    
    dayend = f"E{cells}"
    ws[dayend] = timeout
    currentcell = ws[dayend]
    currentcell.alignment = Alignment(horizontal='right')


# Save the file.
wb.save(filename)

print("Timesheet complete you lazy bastard!")
time.sleep(3)

